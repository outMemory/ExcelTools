# -*- coding: utf-8 -*-
import os
import sys

from PyQt5 import QtCore, QtWidgets
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def reader_excel():
    m = QtWidgets.QFileDialog.getExistingDirectory(
        None, "选取文件夹", "D:/test/待合并表格/")  # 起始路径
    Ui_MainWindow.filepath_excel = m
    print(m)


def save_Excel():
    m = QtWidgets.QFileDialog.getExistingDirectory(
        None, "选取文件夹", "D:/test/合并结果/")  # 起始路径
    Ui_MainWindow.path_save = m
    print(m)


class Ui_MainWindow(object):
    png_ = ''
    list_pdf_1 = []
    list_pdf_2 = []
    list_pdf_1_ = []
    list_pdf_2_ = []
    filepath_excel = 'D:/test/待合并表格'
    path_save = 'D:/test/合并结果'
    biaotou = 1
    message = ''

    def setupUi(self, MainWindow):
        # 主窗口参数设置
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowTitle("拆分表格")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        # 实例化
        MainWindow.setCentralWidget(self.centralwidget)
        MainWindow.setGeometry(QtCore.QRect(550, 250, 500, 500))

        self.fileExcel = QtWidgets.QPushButton(self.centralwidget)
        self.fileExcel.setGeometry(QtCore.QRect(50, 100, 333, 35))
        self.fileExcel.setText("选择要拆分的Excel所在的文件夹")

        self.saveExcel = QtWidgets.QPushButton(self.centralwidget)
        self.saveExcel.setGeometry(QtCore.QRect(50, 160, 333, 35))
        self.saveExcel.setText("选择保存位置")

        # 设置分类依据
        self.exceltitle_text = QtWidgets.QLabel(self.centralwidget)
        self.exceltitle_text.setGeometry(QtCore.QRect(50, 220, 290, 35))
        self.exceltitle_text.setText("根据该列内容进行拆分")
        self.exceltitle = QtWidgets.QLineEdit(self.centralwidget)
        self.exceltitle.setGeometry(QtCore.QRect(350, 220, 33, 35))
        self.exceltitle.setText("1")

        # 设置开始运行按钮
        self.startrun = QtWidgets.QPushButton(self.centralwidget)
        self.startrun.setGeometry(QtCore.QRect(50, 280, 100, 35))
        self.startrun.setText("开始运行")

        # 显示运行信息
        self.yxxx = QtWidgets.QLabel(self.centralwidget)
        self.yxxx.setGeometry(QtCore.QRect(50, 340, 400, 135))
        self.yxxx.setStyleSheet("color:red;font-size:20px")
        self.yxxx.setText("")
        ################button按钮点击事件回调函数################
        self.saveExcel.clicked.connect(save_Excel)
        self.fileExcel.clicked.connect(reader_excel)
        self.startrun.clicked.connect(self.start)

    def start(self):
        try:
            self.biaotou = int(self.exceltitle.text())
        except Exception as e:
            self.message = str(e)
        try:
            for i1_ in os.listdir(self.filepath_excel):
                if i1_.endswith('.xlsx'):
                    self.message = excel(
                        self.filepath_excel + '/' + i1_,
                        self.path_save,
                        self.biaotou)
        except Exception as e:
            self.message = str(e)

        self.yxxx.setStyleSheet("color:red")
        self.yxxx.setStyleSheet("font-size:20px")
        if len(str(self.message)) == 0:
            self.yxxx.setText("搞完了，退出去吧")
        else:
            self.yxxx.setText(str(self.message) + '\n' + '运行结束')
            print(self.message)


def excel(path1, path2, lie):
    """
    根据
    :param path1:
    :param path2:
    :param lie:
    :return:
    """
    try:
        excel_old = load_workbook(path1)
        excel_old_sheet = excel_old.active
        excel_index = []

        # 取第lie行的值形成筛选列表
        for i in range(1, excel_old_sheet.max_row + 1):
            if excel_old_sheet.cell(row=i, column=lie).value is not None:
                excel_index.append(
                    excel_old_sheet.cell(
                        row=i, column=lie).value)
        # 转一下格式
        excel_index = set(excel_index)
        excel_index = list(excel_index)
        # 设置表头
        biaotou = []
        for i_b in range(1, excel_old_sheet.max_column + 1):
            biaotou.append(excel_old_sheet.cell(row=1, column=i_b).value)

        for i_index in range(0, len(excel_index)):
            excel_new = Workbook()
            excel_new_sheet = excel_new.create_sheet('sheet1', index=0)
            excel_new_sheet.append(biaotou)
            for i1 in range(1, excel_old_sheet.max_row + 1):
                excel_newzhi = []
                if excel_old_sheet.cell(
                        row=i1, column=lie).value == excel_index[i_index]:
                    for j1 in range(1, excel_old_sheet.max_column + 1):
                        excel_newzhi.append(
                            excel_old_sheet.cell(
                                row=i1, column=j1).value)
                if len(excel_newzhi) != 0:
                    excel_new_sheet.append(excel_newzhi)
            biaowei = []  # 添加表格末行
            for ii in range(1, excel_old_sheet.max_column + 1):
                abc = get_column_letter(ii)
                biaowei.append('=sum(' + abc + '1:' + abc +
                               str(excel_new_sheet.max_row) + ')')
            excel_new_sheet.append(biaowei)
            excel_new.save(path2 + '/' + str(excel_index[i_index]) + '.xlsx')
            print(path2 + '/' + str(excel_index[i_index]) + '.xlsx')
            print(i_index, '/', len(excel_index))
    except Exception as ee:
        return str(ee)


if __name__ == '__main__':
    # excel('C:/Users/Gis04/Documents/ArcGIS/scratch/工作簿2.xlsx', 'D:/test/拆分结果/qq2', 1)
    app = QtWidgets.QApplication(sys.argv)
    mainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(mainWindow)
    mainWindow.show()
    sys.exit(app.exec_())
