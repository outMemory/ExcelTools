from openpyxl import load_workbook
import os
from openpyxl import Workbook


def del_excel(path, biaotou):
    """

    :param path:
    :param biaotou: 标题所在行
    :return:
    """
    file_list = os.walk('D:/test/待合并表格/')  # 获取这个文件夹下所有的excel文档。
    print(file_list)
    wb = Workbook()
    ws1 = wb.create_sheet('a', index=0)
    # sheetvalue=[]

    # title_list = ['部门', '岗位', '姓名', '日期', '时间类别', '项目合同编号',
    #               '项目负责人', '项目名称', '项目阶段', '工作内容', '专业',
    #               '有效工作时间', '备注']  # sheet表的表头创建
    title_list = []

    for file in file_list:
        path_list = file[2]
        for ph in path_list:  # 迭代遍历所有excel 文档
            ph = ph.replace('~$', '')
            if str(ph).endswith('sx'):
                path = 'D:/test/待合并表格/' + ph
                wb_1 = load_workbook(path)
                wb_1_sheet = wb_1.active
                for i1 in range(1, wb_1_sheet.max_row + 1, 1):
                    hangzhi = []
                    for j1 in range(1, wb_1_sheet.max_column + 1, 1):
                        if j1 <= biaotou:
                            title_list.append(wb_1_sheet.cell(row=i1, column=j1))
                            ws1.append(title_list)
                        zhi = wb_1_sheet.cell(row=i1, column=j1)
                        if zhi is not None:
                            hangzhi.append(zhi.value)
                    ws1.append(hangzhi)
                    # print(zhi)
    wb.save('D:/FX.xlsx')


if __name__ == '__main__':
    del_excel()
