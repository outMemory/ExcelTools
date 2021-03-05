# -*- coding: utf-8 -*-
import os
import sys
import time

from PyQt5 import QtCore, QtWidgets
from openpyxl import Workbook
from openpyxl import load_workbook



class Ui_MainWindow(object):
    png_ = ''
    list_pdf_1 = []  # 不动产权属来源资料
    list_pdf_2 = []  # 登记申请书
    list_pdf_1_ = []
    list_pdf_2_ = []
    filepath_excel = 'D:/test/待合并表格'
    path_save = 'D:/test/合并结果'
    biaotou = 1

    def setupUi(self, MainWindow):
        # 主窗口参数设置
        MainWindow.setObjectName("MainWindow")
        MainWindow.setWindowTitle("刘静怡特定需求合并表格")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")

        # 实例化
        MainWindow.setCentralWidget(self.centralwidget)
        MainWindow.setGeometry(QtCore.QRect(550, 250, 500, 500))

        self.fileExcel = QtWidgets.QPushButton(self.centralwidget)
        self.fileExcel.setGeometry(QtCore.QRect(50, 100, 333, 35))
        self.fileExcel.setText("选择要合并的Excel所在的文件夹")

        self.saveExcel = QtWidgets.QPushButton(self.centralwidget)
        self.saveExcel.setGeometry(QtCore.QRect(50, 160, 333, 35))
        self.saveExcel.setText("选择保存位置")

        # 设置表头
        self.exceltitle_text = QtWidgets.QLabel(self.centralwidget)
        self.exceltitle_text.setGeometry(QtCore.QRect(50, 220, 290, 35))
        self.exceltitle_text.setText("表头所在行数，合并单元格会被拆散")
        self.exceltitle = QtWidgets.QLineEdit(self.centralwidget)
        self.exceltitle.setGeometry(QtCore.QRect(350, 220, 33, 35))
        self.exceltitle.setText("1")

        # 设置开始运行按钮
        self.startrun = QtWidgets.QPushButton(self.centralwidget)
        self.startrun.setGeometry(QtCore.QRect(50, 280, 100, 35))
        self.startrun.setText("开始运行")

        # 显示运行信息
        self.yxxx = QtWidgets.QLabel(self.centralwidget)
        self.yxxx.setGeometry(QtCore.QRect(50, 340, 400, 135))
        self.yxxx.setStyleSheet("color:red;font-size:20px")
        self.yxxx.setText("")

        ################button按钮点击事件回调函数################
        self.saveExcel.clicked.connect(self.save_Excel)
        self.fileExcel.clicked.connect(self.reader_excel)
        self.startrun.clicked.connect(self.start)

    def start(self):
        try:
            self.biaotou = int(self.exceltitle.text())
        except Exception as e:
            print(e)
            time.sleep(10)
            sys.exit(0)
        print(self.biaotou)
        del_excel(self.filepath_excel, self.path_save, self.biaotou)
        print('11111111')
        # 先重置qlable的大小才能刷新显示值
        # self.yxxx.resize(400, 135)
        self.yxxx.setStyleSheet("color:red")
        self.yxxx.setStyleSheet("font-size:20px")
        self.yxxx.setText("搞完了，退出去吧")

    def reader_excel(self):
        m = QtWidgets.QFileDialog.getExistingDirectory(
            None, "选取文件夹", "D:/test/待合并表格/")  # 起始路径
        Ui_MainWindow.filepath_excel = m
        print(m)

    def save_Excel(self):
        m = QtWidgets.QFileDialog.getExistingDirectory(
            None, "选取文件夹", "D:/test/合并结果/")  # 起始路径
        Ui_MainWindow.path_save = m
        print(m)


def del_excel(path1, path2, biaotou):
    """

    :param path1:
    :param path2:
    :param biaotou: 定义表头所在的行数，必须是连续值
    :return:
    """
    file_list = os.walk(path1)  # 获取这个文件夹下所有的excel文档。
    wb = Workbook()
    ws1 = wb.create_sheet('a1', index=0)

    for file in file_list:  # 遍历所有的文件夹
        i = 0  # 控制后续的表头的写入，只有第一个文件才写入表头，后续跳过表头
        path_list = file[2]
        for ph in path_list:  # 迭代遍历所有excel 文档，不能迭代到子文件夹中的文件
            ph = ph.replace('~$', '')
            if str(ph).endswith('sx'):
                path = path1 + '/' + ph
                wb_1 = load_workbook(path)
                wb_1_sheet = wb_1.active
                # 写入表头，缺点在于有多少个表格，就会重复写入相同表头多少次，效率不行
                # 后面的正常写入表格内容时，跳过表头
                print('行数', wb_1_sheet.max_row)
                if i == 0:
                    for ii in range(1, biaotou + 1):
                        title_list = []
                        for jj in range(1, wb_1_sheet.max_column + 1):
                            title_list.append(
                                wb_1_sheet.cell(
                                    row=ii, column=jj).value)
                        ws1.append(title_list)
                    i += 1
                else:
                    pass
                for i1 in range(1 + biaotou, wb_1_sheet.max_row + 1, 1):
                    hangzhi = []
                    i1value = wb_1_sheet.cell(row=i1, column=4).value
                    if i1value is not None:
                        for j1 in range(1, wb_1_sheet.max_column + 1, 1):
                            hangzhi.append(
                                wb_1_sheet.cell(
                                    row=i1, column=j1).value)
                        ws1.append(hangzhi)
            else:
                print(2)
    wb.save(path2 + '/结果.xlsx')


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    mainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(mainWindow)
    mainWindow.show()
    sys.exit(app.exec_())
