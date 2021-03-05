# -*- coding: utf-8 -*-

import openpyxl.workbook.workbook
import openpyxl.worksheet.worksheet
from openpyxl import Workbook
from openpyxl import load_workbook

"""
将输入的excel中的多个子表合并到一个子表中，并默认新增一列用来存放子表的名称
"""


def sheet_meger(
        excel_book: 'openpyxl.workbook.workbook.Workbook',
        new_excel_book_sheet: 'openpyxl.worksheet.worksheet.Worksheet',
        biaotou: 'int'):
    """
    合并指定excel表格中的所有子表，没有做跳过表头的功能
    :param biaotou:
    :param excel_book:
    :param new_excel_book_sheet:
    :return:
    """
    # 先写表头
    if biaotou > 0:
        for i1 in range(1, biaotou + 1):
            biaotou_list = []
            for j1 in range(1,
                            excel_book[excel_book.sheetnames[0]].max_column):
                biaotou_list.append(
                    excel_book[excel_book.sheetnames[0]].cell(row=i1, column=j1).value)
            new_excel_book_sheet.append(biaotou_list)
    # 跳过表头写内容
    for a in excel_book.sheetnames:
        a_sheet = excel_book[a]
        for i in range(1 + biaotou, a_sheet.max_row + 1):
            sheet_cell_list = []
            for j in range(1, a_sheet.max_column + 1):
                sheet_cell_list.append(a_sheet.cell(row=i, column=j).value)
            new_excel_book_sheet.append(sheet_cell_list)


def inser_column(excel_book: 'openpyxl.workbook.workbook.Workbook'):
    """
    在指定列插入当前sheet表的名字
    :return:
    """
    for name in excel_book.sheetnames:
        sheet = excel_book[name]
        for i2 in range(1, sheet.max_row + 1):
            sheet.cell(row=i2, column=25).value = name


if __name__ == '__main__':
    book1 = load_workbook('C:/Users/Gis04/Downloads/2021公务员录取名单.xlsx')
    book2 = Workbook()
    sheet2 = book2.create_sheet('a', index=0)
    inser_column(book1)
    sheet_meger(book1, sheet2, 1)
    book2.save('D:/gwy.xlsx')
