# -*- coding: utf-8 -*-
import os
import sys
import time

from PyQt5 import QtCore, QtWidgets
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, get_column_interval
import openpyxl.worksheet.worksheet
from copy import copy


class excel_tools1:
    # def __init__(self):
    tishi = ''
    savefile = Workbook()
    savefile_sheet = savefile.create_sheet('sheet', index=0)

    @staticmethod
    def save(path):
        excel_tools1.savefile.save(path)

    # @staticmethod  # 静态方法 类或实例均可调用
    def shouhang(
            self,
            excel_sheet1: 'openpyxl.worksheet.worksheet.Worksheet',
            excel_sheet2: 'openpyxl.worksheet.worksheet.Worksheet',
            hang: 'int' = 0, tiaoguo: 'int' = 0):
        """
        会按照原文件名在新的位置生成新的表格，如果原excel表中已经有标题了，那么选hang=0，表示不添加表头
        :param tiaoguo: 设置需要跳过的行数，应对某些表格表头的重复内容
        :param excel_sheet1:
        :param excel_sheet2: 所有excel所在文件夹
        :param hang: 表头所在行
        :return:
        """
        # excel_sheet2 = excelfile2.active
        # 将模板excel中的第一行作为表头写入到path2中的每一个excel表中
        if hang > 0:
            for i in range(1, hang + 1):
                biaotou = []
                for j in range(1, excel_sheet1.max_column + 1):
                    biaotou.append(
                        excel_sheet1.cell(
                            row=i, column=j).value)
                excel_tools1.savefile_sheet.append(biaotou)

        for i1 in range(tiaoguo + 1, excel_sheet2.max_row + 1):
            sheet_othercells = []  # 表中除表头的其他内容
            for j1 in range(1, excel_sheet2.max_column + 1):
                sheet_othercells.append(
                    excel_sheet2.cell(
                        row=i1, column=j1).value)
            excel_tools1.savefile_sheet.append(sheet_othercells)

    @staticmethod
    def excel_meger(excel_sheet1: 'openpyxl.worksheet.worksheet.Worksheet',
                    excel_sheet2: 'openpyxl.worksheet.worksheet.Worksheet'):
        """
        参数传递限制为表格sheet对象
        :param excel_sheet1:模板
        :param excel_sheet2:要改的表
        :return:
        """
        # 储存所有合并单元格起止的list，['A1:A5', 'D10:F13']
        megerlist = str(excel_sheet1.merged_cells).split(' ')
        for i in megerlist:
            excel_sheet2.merge_cells(i)

    @staticmethod
    def copystyle(excel_sheet1: 'openpyxl.worksheet.worksheet.Worksheet',
                  excel_sheet2: 'openpyxl.worksheet.worksheet.Worksheet'):
        """

        :param excel_sheet1: 模板
        :param excel_sheet2: 要改的数据
        :return:
        """
        for i in range(1, excel_sheet1.max_row + 1):
            for j in range(1, excel_sheet1.max_column + 1):
                excel_sheet2.cell(row=i, column=j).number_format = copy(
                    excel_sheet1.cell(row=i, column=i).number_format)


if __name__ == '__main__':
    print("操作excel的首行及末行")
